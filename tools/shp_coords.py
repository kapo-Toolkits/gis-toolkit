# -*- coding: utf-8 -*-
"""წერტილოვანი shapefile-იდან კოორდინატების Excel-ში გატანა და ფორმატირება.

მიეთითება საქაღალდე → იძებნება წერტილოვანი shapefile-ები. ცნობილი სახელის
მსგავსი (Gas_Pipe_Crossing…) ავტომატურად აირჩევა; სხვა შემთხვევაში მომხმარებელი
ირჩევს. UTM ზონა (37/38) ცნობდება .prj-დან (EPSG:32637/32638); თუ ვერ ცნობს —
მომხმარებელი უთითებს. კოორდინატები გაიტანება Excel-ში ორ ბლოკად:

  • A:D — ArcGIS-ის ნედლი ცხრილი (FID, Id, POINT_X, POINT_Y), ხელუხლებელი;
  • E-დან — გასუფთავებული ასლი: არჩევითი ტექსტური „ქუდი“, შემდეგ
    № / X / Y (+ არჩევით „გადაკვეთის კუთხე“ °-ით), ცენტრირებული, all-borders.

მოთხოვნები: geopandas, pyogrio, pyproj, openpyxl.
"""

import os
import re
import sys
import glob

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from tools.base import ToolFrame
from tools.tooltip import add_tip
from tools.translit import lat_to_geo
from tools.clipboard_html import copy_table
from tools.xlsx_format import FileLockedError, save_workbook

# ცნობილი სახელის ნიმუშები, რომლებსაც ავტომატურად ვირჩევთ
AUTO_PATTERNS = ("gas_pipe_crossing", "gas_pipe_protzone_crossing", "crossing")

# ნაგულისხმევი ტექსტური შაბლონები („ქუდები“)
DEFAULT_TEMPLATES = [
    "nakveTis da dacvis zonis sazRvris kveTis koordinatebi",
    "ნაკვეთის და დაცვის ზონის საზღვრის კვეთის კოორდინატები",
    "ნაკვეთის საზღვრის და დაცვის მეორე ზონის კვეთის კოორდინატები",
]

# გარე ტექსტური ფაილი — თითო ხაზზე ერთი ქუდი (UTF-8). მომხმარებელი Notepad-ში
# ამატებს ქართულ ქუდებს (სადაც აკრეფა ნორმალურად მუშაობს) და აქ ჩნდება.
_APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATES_FILE = os.path.join(_APP_DIR, "shp_coords_templates.txt")

# ---- განლაგების მუდმივები (მაგალითის მიხედვით) ----------------------------
# ნედლი ბლოკი A:D. ასლი იწყება G-დან — 2 ცარიელი სვეტი (E, F) დაშორებით,
# რომ დაკოპირებისას ორიგინალს არ დაედოს.
COPY_COL = 7        # G
TITLE_ROW = 6       # ტექსტური ქუდი (იწყება აქედან)
TITLE_ROWS = 3      # ქუდი = 3 რიგის სიმაღლის merge & center ბლოკი
HEADER_ROW_NO_TITLE = 6
TITLE_FONT = "Sylfaen"   # ქუდის ფონტი დეფოლტად
ANGLE_FMT = '0"°"'   # კუთხის სვეტის ფორმატი — მთელი რიცხვი + გრადუსი (წერტილის გარეშე)


RTR = {
    "heading":   {"en": "Shapefile → coordinates (Excel)",
                  "ka": "Shapefile → კოორდინატები (Excel)"},
    "desc":      {"en": "Pick a folder, choose a point shapefile, and export its "
                        "coordinates to a formatted Excel file.",
                  "ka": "მიუთითე საქაღალდე, აირჩიე წერტილოვანი shapefile და გაიტანე "
                        "მისი კოორდინატები დაფორმატებულ Excel ფაილში."},
    "folder":    {"en": "Folder:", "ka": "საქაღალდე:"},
    "browse":    {"en": "Browse…", "ka": "დათვალიერება…"},
    "shp":       {"en": "Point shapefile:", "ka": "წერტილოვანი shapefile:"},
    "rescan":    {"en": "Rescan", "ka": "თავიდან სკანირება"},
    "zone":      {"en": "UTM zone:", "ka": "UTM ზონა:"},
    "zone_auto": {"en": "auto ({z})", "ka": "ავტომატური ({z})"},
    "zone_unknown": {"en": "not detected", "ka": "ვერ ცნობა"},
    "vtype":     {"en": "Values:", "ka": "მნიშვნელობები:"},
    "v_int":     {"en": "Integer", "ka": "მთელი (integer)"},
    "v_dbl":     {"en": "Decimal (double)", "ka": "ათწილადი (double)"},
    "tmpl":      {"en": "Header template:", "ka": "ტექსტური ქუდი:"},
    "tmpl_none": {"en": "(none)", "ka": "(არცერთი)"},
    "tmpl_add":  {"en": "➕ Add…", "ka": "➕ დამატება…"},
    "tmpl_add_q":{"en": "New header template text:", "ka": "ახალი ქუდის ტექსტი:"},
    "dlg_ok":    {"en": "OK", "ka": "დიახ"},
    "dlg_cancel":{"en": "Cancel", "ka": "გაუქმება"},
    "conv_hint": {"en": "Type in Latin — it is converted to Georgian below "
                        "(t→ტ, T→თ, q→ქ, W→ჭ, S→შ, C→ჩ …).",
                  "ka": "აკრიფე ლათინურით — ქვემოთ ქართულად გადაკეთდება "
                        "(t→ტ, T→თ, q→ქ, W→ჭ, S→შ, C→ჩ …)."},
    "conv_geo":  {"en": "Georgian:", "ka": "ქართულად:"},
    "tmpl_file": {"en": "📄 File", "ka": "📄 ფაილი"},
    "tmpl_bad":  {"en": "The text contains “?” (keyboard issue) — not saved. "
                        "Type in Latin (the converter) or edit the file.",
                  "ka": "ტექსტში „?“-ია (აკრეფის ხარვეზი) — არ შეინახა. აკრიფე "
                        "ლათინურით (კონვერტერით) ან ფაილში დაარედაქტირე."},
    "tmpl_builtin": {"en": "Built-in templates can’t be deleted (edit the file for your own).",
                     "ka": "ჩაშენებული ქუდები ვერ იშლება (შენს ქუდებს ფაილში მართავ)."},
    "tip_tmpl_file": {"en": "Edit the header templates in a text file (UTF-8) — "
                            "type Georgian in Notepad, then reopen the dropdown.",
                      "ka": "ქუდები ტექსტურ ფაილში დაარედაქტირე (UTF-8) — ქართული "
                            "Notepad-ში აკრიფე, მერე ჩამოსაშლელი თავიდან გახსენი."},
    "tmpl_del":  {"en": "🗑 Delete", "ka": "🗑 წაშლა"},
    "tmpl_del_q":{"en": "Delete this header template?",
                  "ka": "წავშალო ეს ტექსტური ქუდი?"},
    "angle":     {"en": "Add “crossing angle” column (°)",
                  "ka": "„გადაკვეთის კუთხე“ სვეტი (°)"},
    "angle_hdr": {"en": "gadakveTis kuTxe", "ka": "გადაკვეთის კუთხე"},
    "angle_all": {"en": "same angle for all (optional):",
                  "ka": "ერთი კუთხე ყველასთვის (არჩევით):"},
    "export":    {"en": "Export & format (Excel)", "ka": "ექსპორტი და ფორმატირება (Excel)"},
    "preview":   {"en": "Preview", "ka": "წინასწარ ნახვა"},
    "batch":     {"en": "Batch: export all…", "ka": "Batch: ყველას ექსპორტი…"},
    "batch_dir": {"en": "Choose output folder for batch export",
                  "ka": "აირჩიე batch-ის გამომავალი საქაღალდე"},
    "batch_done":{"en": "Batch done — exported {ok}, skipped {sk}.",
                  "ka": "Batch დასრულდა — გაიტანა {ok}, გამოტოვდა {sk}."},
    "batch_skip":{"en": "skipped (no zone / no points):",
                  "ka": "გამოტოვდა (ზონა/წერტილი არ არის):"},
    "prev_count":{"en": "{n} point(s)", "ka": "{n} წერტილი"},
    "apply_deg": {"en": "Add ° to angle column…", "ka": "° დაუმატე კუთხის სვეტს…"},
    "pick_xlsx": {"en": "Choose the Excel file", "ka": "აირჩიე Excel ფაილი"},
    "deg_done":  {"en": "° applied to {n} angle cell(s).",
                  "ka": "° დაემატა {n} კუთხის უჯრას."},
    "deg_nocol": {"en": "No “crossing angle” column found in the file.",
                  "ka": "ფაილში „გადაკვეთის კუთხე“ სვეტი ვერ მოიძებნა."},
    "no_folder": {"en": "Specify a valid folder.", "ka": "მიუთითე არსებული საქაღალდე."},
    "no_shp":    {"en": "No point shapefiles found in the folder.",
                  "ka": "საქაღალდეში წერტილოვანი shapefile ვერ მოიძებნა."},
    "pick_shp":  {"en": "Choose a point shapefile first.",
                  "ka": "ჯერ აირჩიე წერტილოვანი shapefile."},
    "need_zone": {"en": "Could not detect the UTM zone — please choose 37 or 38.",
                  "ka": "UTM ზონა ვერ ცნობა — აირჩიე 37 ან 38."},
    "no_crs_q":  {"en": "This shapefile has no CRS (.prj). Coordinates will be "
                        "used as-is, treated as UTM zone {z}. Continue?",
                  "ka": "ამ shapefile-ს კოორდინატთა სისტემა (.prj) არ აქვს. "
                        "კოორდინატები გამოყენებული იქნება როგორც არის — UTM ზონა {z}. "
                        "გავაგრძელო?"},
    "no_crs_warn":{"en": "no CRS (.prj) — coordinates used as UTM zone {z}",
                   "ka": "CRS (.prj) არ არის — კოორდინატები UTM ზონა {z}-ად"},
    "no_points": {"en": "The shapefile has no points.",
                  "ka": "shapefile-ში წერტილები არ არის."},
    "found_shp": {"en": "Found {n} point shapefile(s).",
                  "ka": "ნაპოვნია {n} წერტილოვანი shapefile."},
    "detected":  {"en": "Selected: {name} | zone {z} | {n} point(s).",
                  "ka": "არჩეულია: {name} | ზონა {z} | {n} წერტილი."},
    "done":      {"en": "Exported {n} points → {path}",
                  "ka": "გაიტანა {n} წერტილი → {path}"},
    "save_title":{"en": "Save Excel", "ka": "Excel-ის შენახვა"},
    "err":       {"en": "Error", "ka": "შეცდომა"},
    "file_locked":{"en": "Could not save — the file is open in Excel. "
                         "Close it and try again.",
                   "ka": "ვერ შეინახა — ფაილი Excel-ში გახსნილია. "
                         "დახურე და სცადე ხელახლა."},
    "load_err":  {"en": "Could not read the shapefile:",
                  "ka": "shapefile ვერ წაიკითხა:"},

    # tooltip-ები
    "tip_browse": {"en": "Folder to search for point shapefiles.",
                   "ka": "საქაღალდე წერტილოვანი shapefile-ების საძებნელად."},
    "tip_rescan": {"en": "Re-scan the folder for point shapefiles.",
                   "ka": "საქაღალდის ხელახლა სკანირება წერტილოვან shapefile-ებზე."},
    "tip_shp":    {"en": "Choose the point shapefile to export.",
                   "ka": "აირჩიე გასატანი წერტილოვანი shapefile."},
    "tip_tmpl_add": {"en": "Add a new header (title) template.",
                     "ka": "ახალი ტექსტური ქუდის დამატება."},
    "tip_tmpl_del": {"en": "Delete the selected header template.",
                     "ka": "არჩეული ტექსტური ქუდის წაშლა."},
    "tip_angle":  {"en": "Add a “crossing angle” column formatted with ° (degrees).",
                   "ka": "„გადაკვეთის კუთხე“ სვეტი ° (გრადუსი) ფორმატით."},
    "tip_export": {"en": "Export the selected shapefile to a formatted Excel file.",
                   "ka": "არჩეული shapefile-ის ექსპორტი დაფორმატებულ Excel-ში."},
    "clip":       {"en": "📋 Copy", "ka": "📋 კოპირება"},
    "tip_clip":   {"en": "Copy the formatted block (header + bordered №/X/Y) to the "
                        "clipboard — paste directly into Word/Excel.",
                   "ka": "დაფორმატებული ბლოკის (ქუდი + ჩარჩოებიანი №/X/Y) კოპირება "
                        "ბუფერში — Word/Excel-ში პირდაპირ ჩააფეისთე."},
    "clip_done":  {"en": "Copied to clipboard ({n} rows).",
                   "ka": "დაკოპირდა ბუფერში ({n} რიგი)."},
    "clip_fail":  {"en": "Could not copy to clipboard.",
                   "ka": "ბუფერში კოპირება ვერ მოხერხდა."},
    "tip_preview":{"en": "Preview the coordinate table before export.",
                   "ka": "კოორდინატების ცხრილის წინასწარ ნახვა ექსპორტამდე."},
    "tip_batch":  {"en": "Export every point shapefile in the folder at once.",
                   "ka": "საქაღალდის ყველა წერტილოვანი shapefile ერთბაშად."},
    "tip_apply_deg": {"en": "Add ° to all numbers in the angle column of an Excel file.",
                      "ka": "° დაუმატე კუთხის სვეტის ყველა რიცხვს Excel ფაილში."},
}


class ShpCoordsTool(ToolFrame):
    tid = "shp_coords"
    CATALOG = RTR          # tr() მოდის ToolFrame-იდან (საერთო lookup)

    def _state(self):
        return self.app.tool_state.setdefault(self.tid, {})

    # ---- შენახული ტექსტური შაბლონები ----
    @staticmethod
    def _file_templates():
        """გარე ფაილიდან ქუდები. მედეგია კოდირებაზე: utf-8-sig (BOM) → utf-8 →
        cp1251 → cp1252 (Notepad-ის ANSI-ს შემთხვევაშიც წაიკითხავს)."""
        try:
            data = open(TEMPLATES_FILE, "rb").read()
        except OSError:
            return []
        text = None
        for enc in ("utf-8-sig", "utf-8", "cp1251", "cp1252"):
            try:
                text = data.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            return []
        return [ln.strip() for ln in text.splitlines()
                if ln.strip() and not ln.startswith("#")]

    def _templates(self):
        """ნაგულისხმევი (ჩაშენებული) + გარე ფაილი — ერთადერთი წყაროები (დუბლ. გარეშე).
        ერთადერთი რედაქტირებადი — ფაილი (add/delete და Notepad ერთსა და იმავეს ცვლის)."""
        merged = list(DEFAULT_TEMPLATES)
        for t in self._file_templates():
            if t and t not in merged:
                merged.append(t)
        return merged

    def _write_file_templates(self, lines):
        """ფაილის გადაწერა (header + user ქუდები, UTF-8)."""
        try:
            with open(TEMPLATES_FILE, "w", encoding="utf-8") as f:
                f.write("# თითო ხაზზე ერთი ტექსტური ქუდი (UTF-8) / one header per line\n")
                for ln in lines:
                    f.write(ln + "\n")
        except OSError as e:
            messagebox.showerror(self.tr("err"), str(e))

    def _migrate_config_templates(self):
        """ძველი კონფიგში შენახული ქუდები → ფაილში (უვარგისი „?“-იანები იშლება)."""
        cfg = self.app.get_tool_config(self.tid)
        if "templates" not in cfg:
            return
        good = [t for t in (cfg.get("templates") or [])
                if t and "?" not in t and t not in DEFAULT_TEMPLATES]
        if good:
            files = self._file_templates()
            self._write_file_templates(files + [t for t in good if t not in files])
        cfg.pop("templates", None)
        self.app.set_tool_config(self.tid, cfg)

    def _open_templates_file(self):
        """ქუდების ფაილს ხსნის სისტემურ რედაქტორში (Notepad-ში ქართული იწერება)."""
        if not os.path.exists(TEMPLATES_FILE):
            # ფაილში მხოლოდ მომხმარებლის ქუდები (ჩაშენებულები ისედ ჩამონათვალშია)
            self._write_file_templates(self._file_templates())
        try:
            if sys.platform == "win32":
                os.startfile(TEMPLATES_FILE)            # noqa: S606
            elif sys.platform == "darwin":
                __import__("subprocess").run(["open", TEMPLATES_FILE])
            else:
                __import__("subprocess").run(["xdg-open", TEMPLATES_FILE])
        except Exception as e:  # noqa: BLE001
            messagebox.showerror(self.tr("err"), str(e))
            return
        self._refresh_templates()

    def build(self):
        pal = self.app.palette
        st = self._state()
        saved = self.app.get_tool_config(self.tid)
        self._shp_map = {}         # basename -> full path
        self._detected_zone = None

        ttk.Label(self, text=self.tr("heading"),
                  font=("Segoe UI", 13, "bold")).grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 4))
        ttk.Label(self, text=self.tr("desc"), foreground=pal["muted"],
                  wraplength=620, justify="left").grid(
            row=1, column=0, columnspan=3, sticky="w", pady=(0, 10))

        # საქაღალდე
        ttk.Label(self, text=self.tr("folder")).grid(row=2, column=0, sticky="w")
        self.folder_var = tk.StringVar(value=st.get("folder") or saved.get("folder") or "")
        ttk.Entry(self, textvariable=self.folder_var).grid(
            row=3, column=0, columnspan=2, sticky="ew", pady=(2, 8))
        add_tip(ttk.Button(self, text=self.tr("browse"), command=self._pick_folder),
                self.tr("tip_browse")).grid(
            row=3, column=2, sticky="ew", padx=(8, 0), pady=(2, 8))

        # shapefile არჩევა
        ttk.Label(self, text=self.tr("shp")).grid(row=4, column=0, sticky="w")
        self.shp_var = tk.StringVar()
        self.shp_combo = ttk.Combobox(self, textvariable=self.shp_var,
                                      state="readonly")
        self.shp_combo.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(2, 8))
        self.shp_combo.bind("<<ComboboxSelected>>", lambda e: self._on_shp_selected())
        add_tip(self.shp_combo, self.tr("tip_shp"))
        add_tip(ttk.Button(self, text=self.tr("rescan"), command=self._scan),
                self.tr("tip_rescan")).grid(
            row=5, column=2, sticky="ew", padx=(8, 0), pady=(2, 8))

        # UTM ზონა
        zrow = ttk.Frame(self)
        zrow.grid(row=6, column=0, columnspan=3, sticky="w", pady=(0, 6))
        ttk.Label(zrow, text=self.tr("zone")).pack(side="left")
        self.zone_var = tk.StringVar(value=st.get("zone", "auto"))
        self.zone_auto_rb = ttk.Radiobutton(zrow, text=self.tr("zone_auto", z="—"),
                                            variable=self.zone_var, value="auto")
        self.zone_auto_rb.pack(side="left", padx=(6, 8))
        ttk.Radiobutton(zrow, text="37", variable=self.zone_var, value="37").pack(side="left")
        ttk.Radiobutton(zrow, text="38", variable=self.zone_var, value="38").pack(side="left", padx=(6, 0))

        # მნიშვნელობის ტიპი
        vrow = ttk.Frame(self)
        vrow.grid(row=7, column=0, columnspan=3, sticky="w", pady=(0, 6))
        ttk.Label(vrow, text=self.tr("vtype")).pack(side="left")
        self.vtype_var = tk.StringVar(value=st.get("vtype", "double"))
        ttk.Radiobutton(vrow, text=self.tr("v_dbl"), variable=self.vtype_var,
                        value="double").pack(side="left", padx=(6, 8))
        ttk.Radiobutton(vrow, text=self.tr("v_int"), variable=self.vtype_var,
                        value="int").pack(side="left")

        # ტექსტური ქუდი
        trow = ttk.Frame(self)
        trow.grid(row=8, column=0, columnspan=3, sticky="ew", pady=(0, 6))
        ttk.Label(trow, text=self.tr("tmpl")).pack(side="left")
        self.tmpl_var = tk.StringVar(value=st.get("template", self.tr("tmpl_none")))
        # postcommand — ჩამოსაშლელის გახსნისას ფაილიდან ხელახლა იკითხება
        # ქართული ფონტი — ზოგ Windows-ზე combobox-ის ნაგულისხმევი ფონტი ქართულს
        # ვერ აჩვენებს („????“); Sylfaen უჭერს მხარს.
        self.tmpl_combo = ttk.Combobox(trow, textvariable=self.tmpl_var, width=48,
                                       postcommand=self._refresh_templates,
                                       font=("Sylfaen", 11))
        self.tmpl_combo.pack(side="left", fill="x", expand=True, padx=(6, 6))
        self._migrate_config_templates()       # ძველი „????“-ები კონფიგიდან გაქრეს
        self._refresh_templates()
        add_tip(ttk.Button(trow, text=self.tr("tmpl_add"), command=self._add_template),
                self.tr("tip_tmpl_add")).pack(side="left")
        add_tip(ttk.Button(trow, text=self.tr("tmpl_del"), command=self._delete_template),
                self.tr("tip_tmpl_del")).pack(side="left", padx=(4, 0))
        add_tip(ttk.Button(trow, text=self.tr("tmpl_file"), command=self._open_templates_file),
                self.tr("tip_tmpl_file")).pack(side="left", padx=(4, 0))

        # გადაკვეთის კუთხე
        arow = ttk.Frame(self)
        arow.grid(row=9, column=0, columnspan=3, sticky="w", pady=(0, 10))
        self.angle_var = tk.BooleanVar(value=st.get("angle", False))
        add_tip(ttk.Checkbutton(arow, text=self.tr("angle"), variable=self.angle_var),
                self.tr("tip_angle")).pack(side="left")
        ttk.Label(arow, text=self.tr("angle_all")).pack(side="left", padx=(14, 4))
        self.angle_all_var = tk.StringVar(value=st.get("angle_all", ""))
        ttk.Entry(arow, textvariable=self.angle_all_var, width=8).pack(side="left")

        actions = ttk.Frame(self)
        actions.grid(row=10, column=0, columnspan=3, sticky="w")
        add_tip(ttk.Button(actions, text=self.tr("export"), command=self._export),
                self.tr("tip_export")).pack(side="left")
        add_tip(ttk.Button(actions, text=self.tr("clip"), command=self._copy_clipboard),
                self.tr("tip_clip")).pack(side="left", padx=(4, 0))
        add_tip(ttk.Button(actions, text=self.tr("preview"), command=self._preview),
                self.tr("tip_preview")).pack(side="left", padx=(12, 0))
        add_tip(ttk.Button(actions, text=self.tr("batch"), command=self._batch),
                self.tr("tip_batch")).pack(side="left", padx=(4, 0))
        add_tip(ttk.Button(actions, text=self.tr("apply_deg"), command=self._apply_degree),
                self.tr("tip_apply_deg")).pack(side="left", padx=(12, 0))

        # წინასწარი ცხრილი — ექსპორტამდე კოორდინატების გადასამოწმებლად
        prev = ttk.Frame(self)
        prev.grid(row=11, column=0, columnspan=3, sticky="nsew", pady=(10, 0))
        self.prev_count = tk.StringVar(value="")
        ttk.Label(prev, textvariable=self.prev_count,
                  foreground=pal["muted"]).pack(anchor="w")
        self.table = ttk.Treeview(prev, columns=("no", "x", "y"),
                                  show="headings", height=8)
        for c, w in (("no", 50), ("x", 130), ("y", 130)):
            self.table.heading(c, text={"no": "№", "x": "X", "y": "Y"}[c])
            self.table.column(c, width=w, anchor="center")
        vsb = ttk.Scrollbar(prev, orient="vertical", command=self.table.yview)
        self.table.configure(yscrollcommand=vsb.set)
        self.table.pack(side="left", fill="both", expand=True)
        vsb.pack(side="left", fill="y")

        self.columnconfigure(0, weight=1)
        self.rowconfigure(11, weight=1)

        if self.folder_var.get():
            self._scan()

    def save_state(self):
        st = self._state()
        st["folder"] = self.folder_var.get()
        st["zone"] = self.zone_var.get()
        st["vtype"] = self.vtype_var.get()
        st["template"] = self.tmpl_var.get()
        st["angle"] = self.angle_var.get()
        st["angle_all"] = self.angle_all_var.get()

    # ---- შაბლონები ----
    def _refresh_templates(self):
        vals = [self.tr("tmpl_none")] + self._templates()
        self.tmpl_combo["values"] = vals
        if self.tmpl_var.get() not in vals:
            self.tmpl_var.set(vals[0])

    def _ask_text(self, prompt, initial=""):
        """ლათინური→ქართული კონვერტერ-მოდალი: მომხმარებელი აკრეფს ლათინურით
        (ASCII — Tk-ის ??? პრობლემა არ ეხება), ქვემოთ ცოცხლად ჩნდება ქართული,
        და OK-ზე ბრუნდება ქართული ტექსტი."""
        pal = self.app.palette
        dlg = tk.Toplevel(self)
        dlg.title("GIS_BOX")
        dlg.transient(self.winfo_toplevel())
        dlg.resizable(False, False)
        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text=prompt, wraplength=480).pack(anchor="w")
        ttk.Label(frm, text=self.tr("conv_hint"), foreground=pal["muted"],
                  wraplength=480, justify="left").pack(anchor="w", pady=(2, 4))
        lat_var = tk.StringVar(value=initial)
        ent = ttk.Entry(frm, textvariable=lat_var, width=64, font=("Consolas", 11))
        ent.pack(fill="x", pady=(0, 6))
        ent.focus_set()
        ttk.Label(frm, text=self.tr("conv_geo")).pack(anchor="w")
        geo_var = tk.StringVar()
        ttk.Label(frm, textvariable=geo_var, font=("Sylfaen", 13),
                  wraplength=480, justify="left").pack(anchor="w", pady=(0, 10))
        lat_var.trace_add("write", lambda *a: geo_var.set(lat_to_geo(lat_var.get())))
        geo_var.set(lat_to_geo(lat_var.get()))
        out = {"v": None}

        def ok(_e=None):
            out["v"] = geo_var.get().strip()
            dlg.destroy()

        btns = ttk.Frame(frm)
        btns.pack(anchor="e")
        ttk.Button(btns, text=self.tr("dlg_ok"), command=ok).pack(side="left")
        ttk.Button(btns, text=self.tr("dlg_cancel"), command=dlg.destroy).pack(side="left", padx=(6, 0))
        ent.bind("<Return>", ok)
        dlg.bind("<Escape>", lambda e: dlg.destroy())
        dlg.grab_set()
        self.wait_window(dlg)
        return out["v"]

    def _add_template(self):
        txt = self._ask_text(self.tr("tmpl_add_q"))
        if not txt:
            return
        if "?" in txt:                         # აკრეფის ხარვეზი — არ შევინახოთ
            messagebox.showwarning("GIS_BOX", self.tr("tmpl_bad"))
            return
        files = self._file_templates()
        if txt not in files and txt not in DEFAULT_TEMPLATES:
            files.append(txt)
            self._write_file_templates(files)   # ერთადერთი წყარო — ფაილი
        self._refresh_templates()
        self.tmpl_var.set(txt)

    def _delete_template(self):
        cur = self.tmpl_var.get()
        if not cur or cur == self.tr("tmpl_none"):
            return
        files = self._file_templates()
        if cur not in files:                    # ჩაშენებულს ვერ წავშლით
            messagebox.showinfo("GIS_BOX", self.tr("tmpl_builtin"))
            return
        if not messagebox.askyesno("GIS_BOX", f"{self.tr('tmpl_del_q')}\n\n{cur}"):
            return
        files.remove(cur)
        self._write_file_templates(files)
        self._refresh_templates()
        self.tmpl_var.set(self.tr("tmpl_none"))

    # ---- საქაღალდე / სკანირება ----
    def _pick_folder(self):
        d = filedialog.askdirectory(title=self.tr("folder"),
                                    initialdir=self.folder_var.get() or None)
        if d:
            self.folder_var.set(os.path.normpath(d))
            cfg = self.app.get_tool_config(self.tid)
            cfg["folder"] = os.path.normpath(d)
            self.app.set_tool_config(self.tid, cfg)
            self._scan()

    def _scan(self):
        folder = self.folder_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("GIS_BOX", self.tr("no_folder"))
            return
        import pyogrio
        self._shp_map = {}
        for shp in sorted(glob.glob(os.path.join(folder, "*.shp"))):
            try:
                info = pyogrio.read_info(shp)
            except Exception:
                continue
            if "Point" in str(info.get("geometry_type") or ""):
                self._shp_map[os.path.basename(shp)] = shp
        names = list(self._shp_map)
        self.shp_combo["values"] = names
        if not names:
            self.shp_var.set("")
            self.app.log("— " + self.tr("no_shp"))
            messagebox.showinfo("GIS_BOX", self.tr("no_shp"))
            return
        self.app.log("— " + self.tr("found_shp", n=len(names)))
        # ავტომატური არჩევა: ცნობილი ნიმუშის მსგავსი, თუ არა — პირველი
        auto = next((n for n in names
                     if any(p in n.lower() for p in AUTO_PATTERNS)), None)
        self.shp_var.set(auto or names[0])
        self._on_shp_selected()

    def _on_shp_selected(self):
        shp = self._shp_map.get(self.shp_var.get())
        if not shp:
            return
        self._detected_zone = self._detect_zone(shp)
        z = self._detected_zone
        self.zone_auto_rb.config(
            text=self.tr("zone_auto", z=(z if z else self.tr("zone_unknown"))))
        # თუ ავტორეჟიმია და ზონა ვერ ცნო — მომხმარებელს ვთხოვთ არჩევას
        n = self._count_points(shp)
        self.app.log(self.tr("detected", name=self.shp_var.get(),
                             z=(z or "?"), n=n))
        self._preview()          # არჩევისთანავე ცხრილში ჩანს

    # ---- shapefile-ის კითხვა ----
    @staticmethod
    def _detect_zone(shp):
        try:
            import pyogrio, pyproj
            crs = pyogrio.read_info(shp).get("crs")
            if not crs:
                return None
            uz = pyproj.CRS.from_user_input(crs).utm_zone   # მაგ. '37N'
            if uz:
                m = re.match(r"(\d+)", uz)
                if m:
                    return int(m.group(1))
        except Exception:
            pass
        return None

    @staticmethod
    def _count_points(shp):
        try:
            import pyogrio
            return int(pyogrio.read_info(shp).get("features") or 0)
        except Exception:
            return 0

    @staticmethod
    def _crs_missing(shp):
        """True — თუ shapefile-ს კოორდინატთა სისტემა (.prj/CRS) არ აქვს."""
        try:
            import pyogrio
            return not pyogrio.read_info(shp).get("crs")
        except Exception:
            return False

    def _read_points(self, shp, zone):
        """დააბრუნებს [(id, x, y), …] მითითებული ზონის კოორდინატებში."""
        import geopandas as gpd
        gdf = gpd.read_file(shp)
        if gdf.crs is not None:
            try:
                gdf = gdf.to_crs(epsg=32600 + zone)
            except Exception:
                pass
        idcol = next((c for c in gdf.columns if c.lower() == "id"), None)
        pts = []
        for _, row in gdf.iterrows():
            geom = row.geometry
            if geom is None or geom.is_empty:
                continue
            if geom.geom_type == "Point":
                coords = [(geom.x, geom.y)]
            elif geom.geom_type == "MultiPoint":
                coords = [(p.x, p.y) for p in geom.geoms]
            else:
                c = geom.centroid
                coords = [(c.x, c.y)]
            idv = row[idcol] if idcol is not None else 0
            for x, y in coords:
                pts.append((idv, x, y))
        return pts

    def _resolve_zone(self, shp=None):
        """მოქმედი ზონა: ხელით არჩეული (37/38) ან ავტო — .prj-დან ცნობილი
        (batch-ისთვის — კონკრეტული ფაილიდან). აბრუნებს 37/38-ს ან None-ს."""
        if self.zone_var.get() in ("37", "38"):
            return int(self.zone_var.get())
        if shp is not None:
            return self._detect_zone(shp)
        return self._detected_zone

    # ---- წინასწარი ცხრილი ----
    def _fill_table(self, points):
        self.table.delete(*self.table.get_children())
        vtype = self.vtype_var.get()
        for i, (_idv, x, y) in enumerate(points, 1):
            xv = int(round(x)) if vtype == "int" else x
            yv = int(round(y)) if vtype == "int" else y
            self.table.insert("", "end", values=(i, xv, yv))
        self.prev_count.set(self.tr("prev_count", n=len(points)))

    def _preview(self):
        shp = self._shp_map.get(self.shp_var.get())
        if not shp:
            return
        zone = self._resolve_zone(shp)
        if zone not in (37, 38):
            self._fill_table([])
            return
        try:
            points = self._read_points(shp, zone)
        except Exception as e:
            self.app.log(f"⚠ {self.tr('load_err')} {e}")
            self._fill_table([])
            return
        self._fill_table(points)

    # ---- ექსპორტი ----
    def _export(self):
        shp = self._shp_map.get(self.shp_var.get())
        if not shp:
            messagebox.showwarning("GIS_BOX", self.tr("pick_shp"))
            return
        zone = self._resolve_zone(shp)
        if zone not in (37, 38):
            messagebox.showwarning("GIS_BOX", self.tr("need_zone"))
            return
        # CRS-ის გარეშე shapefile — გავაფრთხილოთ, რომ კოორდინატები დაუშვებლად
        # ჩაითვლება არჩეულ ზონად.
        if self._crs_missing(shp):
            if not messagebox.askyesno("GIS_BOX", self.tr("no_crs_q", z=zone)):
                return

        try:
            points = self._read_points(shp, zone)
        except Exception as e:
            messagebox.showerror(self.tr("err"), f"{self.tr('load_err')}\n{e}")
            return
        if not points:
            messagebox.showwarning("GIS_BOX", self.tr("no_points"))
            return

        base = os.path.splitext(os.path.basename(shp))[0]
        path = filedialog.asksaveasfilename(
            title=self.tr("save_title"), defaultextension=".xlsx",
            initialfile=base + ".xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return

        template, angle_on, angle_all, vtype = self._current_settings()
        try:
            self._write_workbook(
                path, sheet_name=os.path.basename(shp), points=points,
                value_type=vtype, template=template,
                angle_on=angle_on, angle_all=angle_all)
        except FileLockedError:
            messagebox.showerror(self.tr("err"), self.tr("file_locked"))
            return
        except Exception as e:
            messagebox.showerror(self.tr("err"), str(e))
            return

        self._last_xlsx = path
        msg = self.tr("done", n=len(points), path=path)
        self.app.log("— " + msg)
        messagebox.showinfo("GIS_BOX", msg)

    def _current_settings(self):
        """მიმდინარე UI პარამეტრები: (template, angle_on, angle_all, vtype)."""
        template = self.tmpl_var.get()
        if template == self.tr("tmpl_none"):
            template = ""
        angle_all = None
        raw = self.angle_all_var.get().strip().replace(",", ".")
        if self.angle_var.get() and raw:
            try:
                angle_all = float(raw)
                if angle_all.is_integer():
                    angle_all = int(angle_all)
            except ValueError:
                angle_all = None
        return template, self.angle_var.get(), angle_all, self.vtype_var.get()

    def _block_rows(self, points, vtype, angle_on, angle_all):
        """(headers, rows) — ზუსტად ის, რაც Excel-ის ასლში (° კუთხეზე)."""
        headers = ["№", "X", "Y"] + ([self.tr("angle_hdr")] if angle_on else [])
        rows = []
        for i, (_idv, x, y) in enumerate(points, 1):
            xv = int(round(x)) if vtype == "int" else float(x)
            yv = int(round(y)) if vtype == "int" else float(y)
            row = [i, xv, yv]
            if angle_on:
                row.append(f"{angle_all}°" if angle_all is not None else "")
            rows.append(row)
        return headers, rows

    def _copy_clipboard(self):
        """შექმნილი ბლოკის (ქუდი + ჩარჩოებიანი №/X/Y) კოპირება ბუფერში —
        HTML-ად (Word/Excel-ში ფორმატით ჩაისმება) + ტექსტად (fallback)."""
        shp = self._shp_map.get(self.shp_var.get())
        if not shp:
            messagebox.showwarning("GIS_BOX", self.tr("pick_shp"))
            return
        zone = self._resolve_zone(shp)
        if zone not in (37, 38):
            messagebox.showwarning("GIS_BOX", self.tr("need_zone"))
            return
        try:
            points = self._read_points(shp, zone)
        except Exception as e:
            messagebox.showerror(self.tr("err"), f"{self.tr('load_err')}\n{e}")
            return
        if not points:
            messagebox.showwarning("GIS_BOX", self.tr("no_points"))
            return

        template, angle_on, angle_all, vtype = self._current_settings()
        headers, rows = self._block_rows(points, vtype, angle_on, angle_all)
        ncols = len(headers)

        def esc(v):
            return (str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

        # უბრალო ცხრილი — ჩარჩოები + ცენტრირება, ფონის/ფერის გარეშე (როგორც არის)
        bd = "border:1px solid #000;"
        ce = "text-align:center;vertical-align:middle;padding:2px 10px;"
        html = ['<table style="border-collapse:collapse;'
                'font-family:Sylfaen,Segoe UI,sans-serif;font-size:11pt;">']
        if template:
            html.append(f'<tr><td colspan="{ncols}" '
                        f'style="{bd}{ce}font-weight:bold;">{esc(template)}</td></tr>')
        html.append("<tr>" + "".join(
            f'<td style="{bd}{ce}font-weight:bold;">{esc(h)}</td>' for h in headers) + "</tr>")
        for row in rows:
            html.append("<tr>" + "".join(
                f'<td style="{bd}{ce}">{esc(c)}</td>' for c in row) + "</tr>")
        html.append("</table>")
        fragment = "".join(html)

        # ტექსტი (TSV) — fallback
        lines = []
        if template:
            lines.append(template)
        lines.append("\t".join(str(h) for h in headers))
        for row in rows:
            lines.append("\t".join(str(c) for c in row))
        text = "\r\n".join(lines)

        if copy_table(fragment, text, widget=self):
            msg = self.tr("clip_done", n=len(rows))
            self.app.log("— " + msg)
            self.app.log(msg)
        else:
            messagebox.showerror(self.tr("err"), self.tr("clip_fail"))

    # ---- Batch: საქაღალდის ყველა წერტილოვანი shapefile ერთბაშად ----
    def _batch(self):
        if not self._shp_map:
            messagebox.showinfo("GIS_BOX", self.tr("no_shp"))
            return
        outdir = filedialog.askdirectory(title=self.tr("batch_dir"))
        if not outdir:
            return
        template, angle_on, angle_all, vtype = self._current_settings()
        ok = skipped = 0
        for name, shp in self._shp_map.items():
            zone = self._resolve_zone(shp)      # ხელით არჩეული ან ფაილიდან ცნობილი
            if zone not in (37, 38):
                self.app.log(f"↷ {self.tr('batch_skip')} {name}")
                skipped += 1
                continue
            if self._crs_missing(shp):          # batch-ში მხოლოდ ვაფრთხილებთ
                self.app.log(f"⚠ {name}: {self.tr('no_crs_warn', z=zone)}")
            try:
                points = self._read_points(shp, zone)
            except Exception as e:
                self.app.log(f"✗ {name}: {e}")
                skipped += 1
                continue
            if not points:
                self.app.log(f"↷ {self.tr('batch_skip')} {name}")
                skipped += 1
                continue
            base = os.path.splitext(name)[0]
            outpath = os.path.join(outdir, base + ".xlsx")
            try:
                self._write_workbook(
                    outpath, sheet_name=name, points=points, value_type=vtype,
                    template=template, angle_on=angle_on, angle_all=angle_all)
                self.app.log(f"✓ {base}.xlsx ({len(points)})")
                self._last_xlsx = outpath
                ok += 1
            except FileLockedError:
                self.app.log(f"✗ {name}: {self.tr('file_locked')}")
                skipped += 1
            except Exception as e:
                self.app.log(f"✗ {name}: {e}")
                skipped += 1
        msg = self.tr("batch_done", ok=ok, sk=skipped)
        self.app.log("— " + msg)
        messagebox.showinfo("GIS_BOX", f"{msg}\n\n{outdir}")

    # ---- ° დამატება არსებულ ფაილში (რიცხვების ჩაწერის შემდეგ) ----
    def _apply_degree(self):
        """გახსნის Excel-ს, იპოვის „გადაკვეთის კუთხე“ სვეტს და ყველა მის
        (რიცხვით) უჯრას მიანიჭებს ° ფორმატს — ე.ი. მომხმარებელი ჯერ ჩაწერს
        რიცხვებს, მერე ერთი ღილაკით ყველას დაუმატებს გრადუსს."""
        from openpyxl import load_workbook

        initial = getattr(self, "_last_xlsx", "") or ""
        path = filedialog.askopenfilename(
            title=self.tr("pick_xlsx"),
            initialdir=os.path.dirname(initial) or None,
            initialfile=os.path.basename(initial),
            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            wb = load_workbook(path)
        except Exception as e:
            messagebox.showerror(self.tr("err"), str(e))
            return

        targets = {"გადაკვეთის კუთხე", "gadakvetis kutxe", "gadakveTis kuTxe"}
        n = 0
        for ws in wb.worksheets:
            # ვიპოვოთ სათაურის უჯრა („…კუთხე“ / „…kuTxe“)
            hdr = None
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and (v.strip() in targets
                                               or "კუთხე" in v or "kutxe" in v.lower()):
                        hdr = cell
                        break
                if hdr:
                    break
            if not hdr:
                continue
            for r in range(hdr.row + 1, ws.max_row + 1):
                cell = ws.cell(row=r, column=hdr.column)
                left = ws.cell(row=r, column=hdr.column - 1).value   # Y სვეტი
                if cell.value is None and left is None:
                    continue
                # ტექსტად ჩაწერილი რიცხვი → რიცხვად
                if isinstance(cell.value, str):
                    s = cell.value.strip().replace("°", "").replace(",", ".")
                    try:
                        f = float(s)
                        cell.value = int(f) if f.is_integer() else f
                    except ValueError:
                        pass
                cell.number_format = ANGLE_FMT
                n += 1
        try:
            save_workbook(wb, path)
        except FileLockedError:
            messagebox.showerror(self.tr("err"), self.tr("file_locked"))
            return
        except Exception as e:
            messagebox.showerror(self.tr("err"), str(e))
            return
        if n == 0:
            messagebox.showinfo("GIS_BOX", self.tr("deg_nocol"))
            return
        self._last_xlsx = path
        msg = self.tr("deg_done", n=n)
        self.app.log("— " + msg)
        messagebox.showinfo("GIS_BOX", msg)

    def _write_workbook(self, path, sheet_name, points, value_type,
                        template, angle_on, angle_all):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
        from tools.xlsx_format import write_block, write_title

        wb = Workbook()
        ws = wb.active
        ws.title = (sheet_name[:31] or "Coordinates")

        bold = Font(bold=True)
        hcenter = Alignment(horizontal="center")

        # --- ნედლი ბლოკი A:D (FID, Id, POINT_X, POINT_Y) — ხელუხლებელი ---
        for c, name in enumerate(["FID", "Id", "POINT_X", "POINT_Y"], 1):
            cell = ws.cell(row=1, column=c, value=name)
            cell.font = bold
            cell.alignment = hcenter
        for i, (idv, x, y) in enumerate(points):
            ws.cell(row=2 + i, column=1, value=i)                       # FID
            ws.cell(row=2 + i, column=2, value=(idv if idv is not None else 0))
            ws.cell(row=2 + i, column=3, value=float(x))               # სრული სიზუსტე
            ws.cell(row=2 + i, column=4, value=float(y))

        # --- გასუფთავებული ასლი (გაზიარებული ფორმატირებით) ---
        col = COPY_COL
        ncols = 4 if angle_on else 3
        # ასლის სვეტების სიგანე („გადაკვეთის კუთხე“ უფრო ფართო).
        copy_widths = [max(4, len(str(len(points)))) + 2, 14, 14] + ([22] if angle_on else [])

        if template:
            end_row = write_title(ws, template, TITLE_ROW, col, ncols,
                                  rows=TITLE_ROWS, font_name=TITLE_FONT,
                                  total_width_chars=sum(copy_widths))
            header_row = end_row + 1                  # 8 + 1 = 9
        else:
            header_row = HEADER_ROW_NO_TITLE

        headers = ["№", "X", "Y"] + ([self.tr("angle_hdr")] if angle_on else [])
        rows = []
        for i, (idv, x, y) in enumerate(points, 1):
            xv = int(round(x)) if value_type == "int" else float(x)
            yv = int(round(y)) if value_type == "int" else float(y)
            rows.append([i, xv, yv] + ([angle_all] if angle_on else []))

        write_block(ws, header_row, col, headers, rows,
                    angle_index=(3 if angle_on else None),
                    angle_fmt=ANGLE_FMT, col_widths=copy_widths)

        # A:D სვეტების სიგანე
        for c in range(1, 5):
            ws.column_dimensions[get_column_letter(c)].width = 14

        ws.freeze_panes = "A2"
        save_workbook(wb, path)
