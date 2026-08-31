# -*- coding: utf-8 -*-
"""ტესტები — რიცხვების გასუფთავება და Excel-ის ფორმატირება (format_xlsx)."""

import openpyxl

from tools.coordextract.export import _clean_number, rows_to_xlsx, format_xlsx


def test_clean_number():
    assert _clean_number("498195") == 498195
    assert _clean_number(" 498 195 ") == 498195
    assert _clean_number("4665437,0") == 4665437
    assert _clean_number("4G98195") == 498195      # OCR-ის ასო რიცხვში
    assert _clean_number("") is None
    assert _clean_number("12.5") == 12.5
    assert _clean_number(42) == 42


def test_format_xlsx_layout(tmp_path):
    p = str(tmp_path / "x.xlsx")
    rows = [("1", 498195, 4665437), ("2", 498266, 4665443)]
    rows_to_xlsx(p, rows)

    saved, n = format_xlsx(p)
    assert saved == p
    assert n == 2

    ws = openpyxl.load_workbook(p).active
    # ორიგინალი A:C ხელუხლებელი
    assert ws["A1"].value == "No"
    assert ws["B2"].value == 498195
    # გასუფთავებული ასლი E-დან, სათაური E6, ნომრები E7/E8
    assert ws["E6"].value == "№"
    assert ws["F6"].value == "X" and ws["G6"].value == "Y"
    assert ws["E7"].value == 1 and ws["E8"].value == 2
    # center + middle + border მონაცემის უჯრაზე
    c = ws["E7"]
    assert c.alignment.horizontal == "center"
    assert c.alignment.vertical == "center"
    assert c.border.left.style == "thin"
