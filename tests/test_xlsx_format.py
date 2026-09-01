# -*- coding: utf-8 -*-
"""ტესტები — გაზიარებული Excel-ფორმატირება (write_block / write_title)."""

import pytest
from openpyxl import Workbook

from tools.xlsx_format import (write_block, write_title, clean_number, DEGREE_FMT,
                               save_workbook, FileLockedError)


def test_write_block_basic():
    wb = Workbook(); ws = wb.active
    rows = [(1, 100, 200), (2, 300, 400)]
    write_block(ws, header_row=1, start_col=1, headers=["№", "X", "Y"], rows=rows,
                col_widths=[6, 12, 12])
    assert ws["A1"].value == "№" and ws["B1"].value == "X"
    assert ws["A2"].value == 1 and ws["C3"].value == 400
    # center + middle + all borders on data
    c = ws["B2"]
    assert c.alignment.horizontal == "center" and c.alignment.vertical == "center"
    for side in (c.border.left, c.border.right, c.border.top, c.border.bottom):
        assert side.style == "thin"
    # column widths applied
    assert ws.column_dimensions["A"].width == 6


def test_write_block_angle_format():
    wb = Workbook(); ws = wb.active
    rows = [(1, 100, 200, 37), (2, 300, 400, None)]
    write_block(ws, header_row=1, start_col=1,
                headers=["№", "X", "Y", "კუთხე"], rows=rows,
                angle_index=3, angle_fmt=DEGREE_FMT)
    assert ws["D2"].number_format == DEGREE_FMT
    assert ws["D3"].number_format == DEGREE_FMT      # ცარიელსაც ედება
    assert ws["D2"].value == 37


def test_write_title_merge_and_border():
    wb = Workbook(); ws = wb.active
    end_row = write_title(ws, "გრძელი ტექსტური ქუდი აქ", top_row=6, start_col=5,
                          ncols=3, rows=3, total_width_chars=34)
    assert end_row == 8
    assert "E6:G8" in [str(m) for m in ws.merged_cells.ranges]
    assert ws["E6"].font.name == "Sylfaen"
    assert ws["E6"].alignment.wrap_text is True
    # outer box complete on all four sides
    assert all(ws.cell(6, c).border.top.style == "thin" for c in (5, 6, 7))
    assert all(ws.cell(8, c).border.bottom.style == "thin" for c in (5, 6, 7))
    assert all(ws.cell(r, 5).border.left.style == "thin" for r in (6, 7, 8))
    assert all(ws.cell(r, 7).border.right.style == "thin" for r in (6, 7, 8))
    # row heights set to fit
    assert ws.row_dimensions[6].height >= 16.0


def test_clean_number_here():
    assert clean_number("4G98195") == 498195
    assert clean_number(None) is None


def test_save_workbook_ok(tmp_path):
    wb = Workbook(); wb.active["A1"] = "x"
    out = tmp_path / "ok.xlsx"
    save_workbook(wb, str(out))
    assert out.exists()


def test_save_workbook_locked_raises():
    class _LockedWB:
        def save(self, path):
            raise PermissionError("file is open")
    with pytest.raises(FileLockedError) as ei:
        save_workbook(_LockedWB(), "busy.xlsx")
    assert ei.value.path == "busy.xlsx"
