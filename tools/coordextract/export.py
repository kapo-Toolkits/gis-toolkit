"""Export coordinate rows to CSV / JSON / XLSX."""

from __future__ import annotations

import csv
import json
from typing import Iterable, List, Sequence, Tuple

# გაზიარებული Excel-ფორმატირება (clean_number/write_block სუფთაა openpyxl-ის
# იმპორტისგან; openpyxl მხოლოდ write_block-ის შიგნით ჩაიტვირთება).
from tools.xlsx_format import clean_number, write_block

_clean_number = clean_number   # backward-compat alias


def format_xlsx(path: str, out_path: str | None = None,
                header: Sequence[str] = ("№", "X", "Y")) -> Tuple[str, int]:
    """Add a cleaned, formatted copy of the coordinate table to an .xlsx.

    The original block (header at A1, data from row 2, columns No/X/Y) is left
    untouched. A tidy copy is written starting two columns to the right of the
    original block and a few rows down (E6 for a 3-column original), with:
      • a numbering column headed “№” holding plain sequential numbers,
      • X/Y values cleaned to plain numbers,
      • all three columns centered (center + middle) with all-borders.

    Returns (saved_path, number_of_rows).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active

    # --- ორიგინალის წაკითხვა: A1 სათაური, მონაცემები მე-2 რიგიდან (No/X/Y) ---
    ORIG_COLS = 3
    data = []
    r = 2
    while True:
        x = ws.cell(row=r, column=2).value
        y = ws.cell(row=r, column=3).value
        if x is None and y is None:          # კოორდინატები აღარ არის — ბოლო
            break
        data.append((x, y))
        r += 1

    # --- ასლის საწყისი პოზიცია: ბოლო სვეტიდან +2 (C→E), რამდენიმე რიგით ქვემოთ ---
    start_col = ORIG_COLS + 2                 # 3 + 2 = 5 (E)
    start_row = 6

    rows = [(i, clean_number(x), clean_number(y))
            for i, (x, y) in enumerate(data, start=1)]

    # სვეტების სიგანე (base + 4)
    wx = max([7] + [len(str(clean_number(x))) for x, _ in data])
    wy = max([7] + [len(str(clean_number(y))) for _, y in data])
    col_widths = [len(str(header[0])) + 4, wx + 4, wy + 4]

    write_block(ws, start_row, start_col, header, rows, col_widths=col_widths)

    save = out_path or path
    wb.save(save)
    return save, len(data)


def _as_number(v):
    """Return an int/float if the value looks numeric, else the original string."""
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", ".")
    if s == "":
        return ""
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return v


def rows_to_csv(path: str, rows: Sequence[Tuple[str, object, object]],
                header: Sequence[str] = ("No", "X", "Y")) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        for r in rows:
            w.writerow(r)


def rows_to_json(path: str, rows: Sequence[Tuple[str, object, object]],
                 keys: Sequence[str] = ("no", "x", "y"),
                 extra: dict | None = None) -> None:
    data = {
        "points": [dict(zip(keys, r)) for r in rows],
    }
    if extra:
        data.update(extra)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def rows_to_xlsx(path: str, rows: Sequence[Tuple[str, object, object]],
                 header: Sequence[str] = ("No", "X", "Y"),
                 sheet_name: str = "Coordinates",
                 note: str | None = None) -> None:
    """Write rows to a modern .xlsx workbook (numbers stored as numbers)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Coordinates"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    for c, name in enumerate(header, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = bold
        cell.alignment = center

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            out = _as_number(val) if c >= 2 else val  # keep label column as-is
            ws.cell(row=r, column=c, value=out)

    for c, name in enumerate(header, 1):
        width = max(len(str(name)), *(len(str(row[c - 1])) for row in rows)) if rows else len(str(name))
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = width + 4

    if note:
        ws.cell(row=len(rows) + 3, column=1, value=note)

    ws.freeze_panes = "A2"
    wb.save(path)
