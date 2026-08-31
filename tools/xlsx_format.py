# -*- coding: utf-8 -*-
"""გაზიარებული Excel-ფორმატირება — „დაფორმატებული კოორდინატების ბლოკი“.

აქ თავმოყრილია სტილი და ლოგიკა, რომელსაც იყენებენ ორივე ხელსაწყო:
``tools.coordextract.export.format_xlsx`` და ``tools.shp_coords`` — ერთგვაროვანი
გამოსავალი, ნაკლები დუბლირება.

``clean_number`` და ``fit_title_rows`` სუფთაა (openpyxl არ სჭირდება); ხოლო
``write_block``/``write_title`` openpyxl-ს ფუნქციის შიგნით აიმპორტებენ, ასე რომ
ამ მოდულის იმპორტი openpyxl-ის გარეშეც შესაძლებელია.
"""

import re
import math

DEGREE_FMT = '0"°"'   # კუთხის ფორმატი — მთელი რიცხვი + გრადუსი (წერტილის გარეშე)


def clean_number(v):
    """მოსახერხებელ რიცხვად გადაქცევა, შეცდომებისგან თავისუფალი.

    სფეისი/ასო/სიმბოლო (OCR ნაგავი) იშლება, მძიმე → ათწილადი წერტილი, მთელი
    რიცხვი int-ად ბრუნდება. None თუ არაფერი რიცხვითი დარჩა."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v) if float(v).is_integer() else v
    s = str(v).strip().replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)          # სფეისი/ასო/სიმბოლო მოშორება
    if s in ("", "-", ".", "-."):
        return None
    try:
        f = float(s)
    except ValueError:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        if not m:
            return None
        f = float(m.group())
    return int(f) if f.is_integer() else f


def fit_title_rows(ws, text, first_row, n_rows, total_width_chars):
    """ქუდის რიგების სიმაღლე ტექსტის სიგრძის მიხედვით — რომ wrap-ული ტექსტი
    merge-ბლოკში ნორმალურად ჩაჯდეს (merge-ს Excel ავტომატურად არ უსწორებს)."""
    per_line = max(10, int(total_width_chars) - 2)     # დაახლ. სიმბოლო/ხაზზე
    lines = max(1, math.ceil(len(str(text)) / per_line))
    needed = lines * 15.0 + 8                           # საჭირო სიმაღლე (pt)
    per_row = max(16.0, needed / n_rows)
    for r in range(first_row, first_row + n_rows):
        ws.row_dimensions[r].height = per_row


def _styles():
    from openpyxl.styles import Alignment, Font, Border, Side
    thin = Side(style="thin", color="000000")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center"),
        "wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "bold": Font(bold=True),
    }


def write_block(ws, header_row, start_col, headers, rows,
                angle_index=None, angle_fmt=DEGREE_FMT, col_widths=None):
    """დაფორმატებული ბლოკი: სათაური + მონაცემები, center+middle და all-borders.

    ``headers`` — სვეტების სახელები (მაგ. № / X / Y [/ კუთხე]).
    ``rows``    — მნიშვნელობების tuple-ების სია (უკვე საბოლოო ფორმაში).
    ``angle_index`` — ინდექსი (ბლოკში) სვეტისა, რომელსაც კუთხის ° ფორმატი დაედება.
    ``col_widths``  — სვეტების სიგანეები (არჩევით)."""
    from openpyxl.utils import get_column_letter
    s = _styles()
    for j, name in enumerate(headers):
        c = ws.cell(row=header_row, column=start_col + j, value=name)
        c.font = s["bold"]
        c.alignment = s["center"]
        c.border = s["border"]
    for i, row in enumerate(rows, start=1):
        rr = header_row + i
        for j, val in enumerate(row):
            c = ws.cell(row=rr, column=start_col + j, value=val)
            c.alignment = s["center"]
            c.border = s["border"]
            if angle_index is not None and j == angle_index:
                c.number_format = angle_fmt
    if col_widths:
        for j, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(start_col + j)].width = w
    return header_row


def write_title(ws, title, top_row, start_col, ncols, rows=3,
                font_name="Sylfaen", size=11, total_width_chars=None):
    """ტექსტური ქუდი — merge & center ბლოკი, wrap text, all-borders და (არჩევით)
    ტექსტის სიგრძეზე მორგებული რიგის სიმაღლე. აბრუნებს ბოლო რიგს."""
    from openpyxl.styles import Alignment, Font
    s = _styles()
    end_row = top_row + rows - 1
    end_col = start_col + ncols - 1
    ws.merge_cells(start_row=top_row, start_column=start_col,
                   end_row=end_row, end_column=end_col)
    t = ws.cell(row=top_row, column=start_col, value=title)
    t.font = Font(name=font_name, bold=True, size=size)
    t.alignment = s["wrap"]
    # all borders — merge-ის შემდეგ დიაპაზონის იტერაციით (გარე ჩარჩო შენარჩუნდება)
    for rr in range(top_row, end_row + 1):
        for cc in range(start_col, end_col + 1):
            ws.cell(row=rr, column=cc).border = s["border"]
    if total_width_chars:
        fit_title_rows(ws, title, top_row, rows, total_width_chars)
    return end_row
